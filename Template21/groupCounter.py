# this is a class of group counter 
# it can be used to read the group contribution vector from the gjf or log file
# the group contribution vector consists of the number of indivisual groups and relative scalar of interaction betweenn groups

# library import area
import os
import re
import openpyxl

import chem

# constants



# units:
# ZPE: kcal/mol
# rotConsts: cm-1
# frequency: cm-1
# hessian: Hartree/Bohr2
# MW: amu
# exponentialDown: cm-1
# hinderedRotorQM1D: angle: degree energy: cm-1
class groupCounter:
	#definetion of comparing pattern
	pattern_gjfCommand = re.compile('^.*#p?.*$')
	pattern_gjfMulti = re.compile('^.*([0-9]+) +([0-9]+).*$')
	pattern_blankLine = re.compile('^ *$')

	# definition of variables
	name = ''
	multi = 0
	geom = []
	formula = ''
	atomsNum = 0

	groupLib = []

	# the default geom and connectivity are from gjf file. Thus self.geom is in gjf format
	# self.groupLib include groupA, groupB and groupA-groupB	
	def __init__(self):
		self.name = ''
		self.multi = 0
		self.geom = []
		self.formula = ''
		self.atomsNum = 0

		self.groupLib = []

	# fileName is the name of the input gjf file
	# directory containing this gjf file. it is needed in case that the gjf file is not in the same directory as this .py code
	# moleculeLabel is the unique label or name of this molecule for convenient reference, such as ethane or pentane
	# if moleculeLabel is not given, the name of molecule would be regarded as the same as the formula
	# the output parameter multi is the multiplicity of the molecule, which is not needed in this group contribution method. Thus even a wrong multiplicity could work. When multi is not given, it would be regarded as 0 as default.
	def readGjfFile(self, fileName, directory='', moleculeLabel=''):
		#definition of flags
		gjfCommand_done = -1
		gjfMulti_done = -1
		geomDone = -1

		#definition of temporary variables
		tmp_m = []		#match result 
		lineStart = 0
		lineEnd = 0

		print fileName
		gjfFile = file(os.path.join(directory, fileName), 'r')
		tmp_lines = gjfFile.readlines()
		for (lineNum, tmp_line) in enumerate(tmp_lines):
			if gjfCommand_done != 1:
				tmp_m = self.pattern_gjfCommand.match(tmp_line)
				if tmp_m:
					gjfCommand_done = 1
			elif gjfMulti_done != 1:
				tmp_m = self.pattern_gjfMulti.match(tmp_line)
				if tmp_m:
					lineStart = lineNum
					self.multi = int(tmp_m.group(2))
					geomDone = 0
					gjfMulti_done = 1
			elif geomDone != 1:
				tmp_m = self.pattern_blankLine.match(tmp_line)
				if tmp_m:
					lineEnd = lineNum
					self.geom = tmp_lines[lineStart+1: lineEnd]
					geomDone = 1


		if geomDone != 1:
			print 'Sorry! The input file is not a standard gjf file!'
		else:
			print 'Gjf file read in successfully!'

		gjfFile.close()

		tmp_mole = chem.molecule()
		tmp_mole.getGjfGeom(self.geom)
		tmp_mole.calcFormula()
		self.formula = tmp_mole.formula
		self.atomsNum = tmp_mole.getAtomsNum()

		if moleculeLabel == '':
			self.name = formula
		else:
			self.name = moleculeLabel

		return self.name, self.multi, self.geom, self.formula, self.atomsNum

	# moleculeLabel is the unique label or name of this molecule for convenient reference, such as ethane or pentane
	# if moleculeLabel is not given, the name of molecule would be regarded as the same as the formula
	# the output parameter multi is the multiplicity of the molecule, which is not needed in this group contribution method. Thus even a wrong multiplicity could work. When multi is not given, it would be regarded as 0 as default.
	def readGjfGeom(self, gjfGeom, moleculeLabel=''):
		self.multi = 0
		self.geom = gjfGeom.strip()

		#definition of flags
		gjfCommand_done = -1
		gjfMulti_done = -1
		geomDone = -1

		#definition of temporary variables
		tmp_m = []		#match result 
		lineStart = 0
		lineEnd = 0

		tmp_mole = chem.molecule()
		tmp_mole.getGjfGeom(geom)
		tmp_mole.calcFormula()
		self.formula = tmp_mole.formula
		self.atomsNum = tmp_mole.getAtomsNum()

		if moleculeLabel == '':
			self.name = formula
		else:
			self.name = moleculeLabel

		return self.name, self.multi, self.geom, self.formula, self.atomsNum

	# this function is used read the template containing the names of groups and group interactions in current database
	# use this function once before writeDBGCVector, because the dimension of the GBGC vector should be determined in advance
	# the default name of the template file is groupTemplate.xlsx
	def readGroupTemplate(self, fileName='groupTemplate.xlsx'):
		if not os.path.exists(fileName):
			print 'Error! Group template file ' + fileName + 'does not exist!'
		
		wbr = openpyxl.load_workbook(fileName)
		shr = wbr.get_sheet_by_name('inputVectors') 
		vectorDimension = int(shr.cell(row=2, column=4).value)
		row_group = 3
		column_group = 6
		all_groups = []
		while True:
			tmp_group = shr.cell(row=row_group, column=column_group).value
			if tmp_group != None and tmp_group != '':
				all_groups.append(tmp_group)
				column_group += 1
			else:
				break
		if vectorDimension != len(all_groups):
			print vectorDimension
			print all_groups
			print 'Error! The dimension of vector is not the same as the length of group vector in template file ' + fileName
		else:
			self.groupLib = all_groups
			print 'Read group template file successfully'
		
		return self.groupLib

	def writeDBGCVector(self, fileName='DBGCVectors.xlsx', overwrite=True):
		if not os.path.exists('DBGCVectors'):
			os.mkdir('DBGCVectors')
		os.chdir('DBGCVectors')

		vectorDimension = len(self.groupLib)
		if overwrite == False and os.path.exists(fileName):
			wbw = openpyxl.load_workbook(fileName)
			shw = wbw.get_sheet_by_name('inputVectors')
			speciesNumber = shw.cell(row=2, column=2).value
		else:
			speciesNumber = 0
			wbw = openpyxl.Workbook()
			shw = wbw.active
			shw.title = 'inputVectors'
			tmp_row = 1
			tmp_col = 1
			shw.cell(row=tmp_row, column=tmp_col).value = 'ID'
			shw.cell(row=tmp_row, column=tmp_col+1).value = 'Number of species'
			shw.cell(row=tmp_row, column=tmp_col+3).value = 'TotalDimesionNumber'
			shw.cell(row=tmp_row, column=tmp_col+5).value = 'DimensionIndex'
			tmp_row = 2
			shw.cell(row=tmp_row, column=tmp_col+3).value = vectorDimension		
			tmp_col = 6
			for i in xrange(vectorDimension):
				shw.cell(row=tmp_row, column=tmp_col+i).value = i+1
				shw.cell(row=tmp_row+1, column=tmp_col+i).value = self.groupLib[i]
			tmp_row = 3
			shw.cell(row=tmp_row, column=vectorDimension+8).value = 'Name'
			shw.cell(row=tmp_row, column=vectorDimension+9).value = 'ReferenceEnergy'

		groupIndex = {}
		tmp_col = 6
		for i in xrange(vectorDimension):
			groupIndex[self.groupLib[i]] = tmp_col+i

		tmp_mole = chem.molecule()
		tmp_mole.getGjfGeom(self.geom)
		tmp_mole.setLabel(self.name)
		tmp_mole.fulfillBonds()
		tmp_groups = tmp_mole.get1stOrderGroup()
		if not set(tmp_groups).issubset(set(self.groupLib)):
			print 'Error! The input molecule does not belong to the family of trained molecules.'
		tmp_groupVector = tmp_mole.getGroupVector6()
		tmp_row = 4+speciesNumber
		tmp_col = 1
		speciesNumber += 1
		shw.cell(row=tmp_row, column=tmp_col).value = speciesNumber
		for i in xrange(vectorDimension):
			shw.cell(row=tmp_row, column=tmp_col+5+i).value = 0.0
		for tmp_vectorEle in tmp_groupVector.keys():
			shw.cell(row=tmp_row, column=groupIndex[tmp_vectorEle]).value = tmp_groupVector[tmp_vectorEle]
		shw.cell(row=tmp_row, column=vectorDimension+8).value = tmp_mole.label

		tmp_row = 2
		tmp_col = 2
		shw.cell(row=tmp_row, column=tmp_col).value = speciesNumber

		wbw.save(fileName)
		os.chdir('../')
		print 'Write group vector successfully!'