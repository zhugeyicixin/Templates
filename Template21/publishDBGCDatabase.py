import openpyxl
import re

# return the indice and the sorted list
def natural_sort(l): 
	convert = lambda text: int(text) if text.isdigit() else text.lower() 
	alphanum_key = lambda key: [ convert(c) for c in re.split('([0-9]+)', key[1]) ] 
	return sorted(enumerate(l), key = alphanum_key)

# input variables
databaseFile = 'database.xlsx'
inputFile = 'inputFile_m6.xlsx'

# variables
all_labels = []
all_formulas = []
all_atomNums = []
all_SCFInFreq = []
all_ZPEInFreq = []
all_enthalpyInFreq = []
all_SPInEnergy = []
all_EnergyWithZPE = []
all_EnthalpyWithZPE = []
all_multi = []
all_geoms = []
all_freqs = []
all_inputLabels = []
all_groupVectors = []
all_enthalpies = []
groupLib = []

# read database excel
wbr1 = openpyxl.load_workbook(databaseFile)
shr1 = wbr1.get_sheet_by_name('speciesInfo')
loaded = False
tmp_row = 2
tmp_col = 1
while not loaded:
	tmp_label = shr1.cell(row=tmp_row, column=tmp_col).value
	if tmp_label == '' or tmp_label == None:
		loaded = True
	else:
		all_labels.append(tmp_label)
		all_formulas.append(shr1.cell(row=tmp_row, column=3).value)
		all_atomNums.append(shr1.cell(row=tmp_row, column=4).value)
		all_SCFInFreq.append(shr1.cell(row=tmp_row, column=5).value)
		all_ZPEInFreq.append(shr1.cell(row=tmp_row, column=6).value)
		all_enthalpyInFreq.append(shr1.cell(row=tmp_row, column=7).value)
		all_SPInEnergy.append(shr1.cell(row=tmp_row, column=8).value)
		all_EnergyWithZPE.append(shr1.cell(row=tmp_row, column=9).value)
		all_EnthalpyWithZPE.append(shr1.cell(row=tmp_row, column=10).value)
		all_multi.append(shr1.cell(row=tmp_row, column=11).value)
		all_geoms.append(shr1.cell(row=tmp_row, column=13).value)
		all_freqs.append(shr1.cell(row=tmp_row, column=14).value)
		tmp_row += 1

# read inputFile excel
wbr2 = openpyxl.load_workbook(inputFile)
shr2 = wbr2.get_sheet_by_name('acyclic')
loaded = False

tmp_row = 3
tmp_col = 6
for i in xrange(170):
	groupLib.append(shr2.cell(row=tmp_row, column=tmp_col+i).value)

tmp_row = 4
tmp_col = 178
while not loaded:
	tmp_label = shr2.cell(row=tmp_row, column=tmp_col).value
	if tmp_label == '' or tmp_label == None:
		loaded = True
	else:
		all_inputLabels.append(tmp_label)
		all_enthalpies.append(shr2.cell(row=tmp_row, column=tmp_col+1).value)
		tmp_vector = []
		for i in xrange(6,176):
			tmp_vector.append(shr2.cell(row=tmp_row, column=i).value)
		all_groupVectors.append(tmp_vector)
		tmp_row += 1

# ranking with the natural order
sorted_inputLabels = natural_sort(all_inputLabels)

# hash the all_inputLabels variable for quick searching
dict_all_labels = dict()
for (index, tmp_label) in enumerate(all_labels):
	dict_all_labels[tmp_label] = index

# write into the database to be published
wbw1 = openpyxl.Workbook()
shw1 = wbw1.active
shw1.title = 'speciesInfo'

tmp_row = 1
tmp_col = 1
shw1.cell(row=tmp_row, column=tmp_col).value = 'ID'
shw1.cell(row=tmp_row, column=tmp_col+1).value = 'SMILES'
shw1.cell(row=tmp_row, column=tmp_col+2).value = 'Formula'
shw1.cell(row=tmp_row, column=tmp_col+3).value = 'Atoms Number'
shw1.cell(row=tmp_row, column=tmp_col+4).value = 'SCF Energy in folder freq (unit: hartree)'
shw1.cell(row=tmp_row, column=tmp_col+5).value = 'ZPE (0 K) Energy in folder freq (unit: hartree)'
shw1.cell(row=tmp_row, column=tmp_col+6).value = 'Enthalpy (298.15 K) in folder freq (unit: hartree)'
shw1.cell(row=tmp_row, column=tmp_col+7).value = 'SP Energy in folder energy (unit: hartree)'
shw1.cell(row=tmp_row, column=tmp_col+8).value = 'SP Energy (0 K) in folder energy corrected with freq scaling factor (unit: hartree)'
shw1.cell(row=tmp_row, column=tmp_col+9).value = 'SP Enthalpy (298.15 K) in folder energy with ZPE and enthalpy correction (unit: hartree)'
shw1.cell(row=tmp_row, column=tmp_col+10).value = 'Calculated standard formation of enthalpy at M062X/def2TZVP//B3LYP/6-31G(d) level of theory (unit: kcal/mol)'
shw1.cell(row=tmp_row, column=tmp_col+11).value = 'Multiplicity (2*S+1)'
shw1.cell(row=tmp_row, column=tmp_col+12).value = 'Connectivity (.gjf format)'
shw1.cell(row=tmp_row, column=tmp_col+13).value = 'Geometry'
shw1.cell(row=tmp_row, column=tmp_col+14).value = 'Frequencies'
shw1.cell(row=tmp_row, column=tmp_col+15).value = 'Group contribution vectors'

tmp_row = 2
tmp_col = 16
for (index, tmp_group) in enumerate(groupLib):
	shw1.cell(row=tmp_row, column=tmp_col+index).value = index+1
	shw1.cell(row=tmp_row+1, column=tmp_col+index).value = tmp_group

tmp_row = 4
tmp_col = 1
for tmp_mole in sorted_inputLabels:
	# tmp_mole is like [index, tmp_label]
	shw1.cell(row=tmp_row, column=tmp_col).value = tmp_mole[1]
	indexInDict_all_labels = dict_all_labels[tmp_mole[1]]
	shw1.cell(row=tmp_row, column=tmp_col+2).value = all_formulas[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+3).value = all_atomNums[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+4).value = all_SCFInFreq[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+5).value = all_ZPEInFreq[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+6).value = all_enthalpyInFreq[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+7).value = all_SPInEnergy[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+8).value = all_EnergyWithZPE[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+9).value = all_EnthalpyWithZPE[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+10).value = all_enthalpies[tmp_mole[0]]
	shw1.cell(row=tmp_row, column=tmp_col+11).value = all_multi[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+13).value = all_geoms[indexInDict_all_labels]
	shw1.cell(row=tmp_row, column=tmp_col+14).value = all_freqs[indexInDict_all_labels]
	for (Misaka, tmp_element) in enumerate(all_groupVectors[tmp_mole[0]]):
		shw1.cell(row=tmp_row, column=tmp_col+15+Misaka).value = tmp_element
	tmp_row += 1
wbw1.save('currentDatabase.xlsx')


