# this is a script to generate smiles from geometry coordinates 

# library import area
import os
import re
import openpyxl
import numpy as np

import chem

# input variables
dataFile = 'database.xlsx'

# variables
all_geoms = []
all_labels = []
resonanceWarning = []
smilesDict = {}

# constants

# read database excel
wbrw = openpyxl.load_workbook(dataFile)
shrw = wbrw.get_sheet_by_name('speciesInfo')
loaded = False
tmp_row = 2
tmp_col = 14
while not loaded:
	tmp_geom = shrw.cell(row=tmp_row, column=tmp_col).value
	if tmp_geom == '' or tmp_geom == None: 
		loaded = True
	else:
		all_geoms.append(tmp_geom)
		all_labels.append(shrw.cell(row=tmp_row, column=1).value)
		tmp_row += 1

# write .sdf file with connectivity and write gjf connectivity into database 

tmp_mole = chem.molecule()
for (index_mole, tmp_geom) in enumerate(all_geoms):
	# resonance structure is the radical like CH2=CHCH2. Only radical for one double bond is considered currently. Thus the accumulated number of resonance for a radical can only be 0 or 2.
	# in the .sdf file, which is used to generate smiles, resonance connectivity is described with bond order 2 and 1. Because bond order 1.5 is not supported in non-aromatic smiles.
	# in the .gjf connectivity, which is written into the database, the bond order 1.5 is used.
	resonance = 0
	mole_geom = tmp_geom.strip()
	mole_geom = mole_geom.split('\n')
	tmp_mole.getGjfGeom(mole_geom)
	tmp_mole.calcFormula()
	tmp_mole.fulfillBonds()
	tmp_mole.setLabel(all_labels[index_mole])

	# write .sdf file with connectivity
	fw = file(tmp_mole.label+'.sdf', 'w')
	fw.write(tmp_mole.label + '''

generated based on coordinates by chem.py
''' + '%3d'%tmp_mole.getAtomsNum() + '%3d'%len(tmp_mole.bonds) + '  0  0  0  0  0  0  0  0999 V2000\n')
	for tmp_atom in tmp_mole.atoms:
		fw.write('%10.4f'%tmp_atom.coordinate[0] + '%10.4f'%tmp_atom.coordinate[1] + '%10.4f'%tmp_atom.coordinate[2] + '%2s'%tmp_atom.symbol + '   0  0  0  0  0  0  0  0  0  0  0  0\n')
	for tmp_atom in tmp_mole.atoms:
		tmp_bonds = {}
		for (index_child, tmp_child) in enumerate(tmp_atom.children):
			if tmp_child.label > tmp_atom.label:
				tmp_bondOrder = tmp_atom.bonds[index_child].bondOrder
				if tmp_bondOrder - int(tmp_bondOrder) > 1e-3:
					resonance += 1
					if resonance == 1:
						tmp_bondOrder -= 0.5
					elif resonance == 2:
						tmp_bondOrder += 0.5	 
					else:
						print 'Error! The variable resonance is larger than 2!'
				if np.abs(tmp_bondOrder - round(tmp_bondOrder)) > 1e-3:
					print 'Error! The bond order used to generate smiles is not integer!'
				tmp_bonds[tmp_child.label] = int(round(tmp_bondOrder))
		for tmp_label in sorted(tmp_bonds.keys()):
			fw.write('%3d'%tmp_atom.label + '%3d'%tmp_label + '%3d'%int(tmp_bonds[tmp_label]) + '  0  0  0  0\n')

	fw.write('''M  END
$$$$''')
	fw.close()

	if resonance != 0:
		print 'Attention! This radical is a resonance structure! ' + tmp_mole.label
		resonance = 0
		fw = file(tmp_mole.label+'_resonance.sdf', 'w')
		fw.write(tmp_mole.label + '''

generated based on coordinates by chem.py
''' + '%3d'%tmp_mole.getAtomsNum() + '%3d'%len(tmp_mole.bonds) + '  0  0  0  0  0  0  0  0999 V2000\n')
		for tmp_atom in tmp_mole.atoms:
			fw.write('%10.4f'%tmp_atom.coordinate[0] + '%10.4f'%tmp_atom.coordinate[1] + '%10.4f'%tmp_atom.coordinate[2] + '%2s'%tmp_atom.symbol + '   0  0  0  0  0  0  0  0  0  0  0  0\n')
		for tmp_atom in tmp_mole.atoms:
			tmp_bonds = {}
			for (index_child, tmp_child) in enumerate(tmp_atom.children):
				if tmp_child.label > tmp_atom.label:
					tmp_bondOrder = tmp_atom.bonds[index_child].bondOrder
					if tmp_bondOrder - int(tmp_bondOrder) > 1e-3:
						resonance += 1
						if resonance == 1:
							tmp_bondOrder += 0.5
						elif resonance == 2:
							tmp_bondOrder -= 0.5	 
						else:
							print 'Error! The variable resonance is larger than 2!'
					if np.abs(tmp_bondOrder - round(tmp_bondOrder)) > 1e-3:
						print 'Error! The bond order used to generate smiles is not integer!'
					tmp_bonds[tmp_child.label] = int(round(tmp_bondOrder))
			for tmp_label in sorted(tmp_bonds.keys()):
				fw.write('%3d'%tmp_atom.label + '%3d'%tmp_label + '%3d'%int(tmp_bonds[tmp_label]) + '  0  0  0  0\n')

		fw.write('''M  END
$$$$''')
		fw.close()

	# write .gjf connectivity into database
	tmp_connectivity = ''
	for tmp_atom in tmp_mole.atoms:
		tmp_bonds = {}
		for (index_child, tmp_child) in enumerate(tmp_atom.children):
			if tmp_child.label > tmp_atom.label:
				tmp_bondOrder = tmp_atom.bonds[index_child].bondOrder
				tmp_bonds[tmp_child.label] = tmp_bondOrder
		tmp_connectivity += ' ' + str(tmp_atom.label) 
		for tmp_label in sorted(tmp_bonds.keys()):
			tmp_connectivity += ' ' + str(tmp_label) + ' %.1f'%tmp_bonds[tmp_label]
		tmp_connectivity += '\n'
	shrw.cell(row=2+index_mole, column=15).value = tmp_connectivity

# openbabel convert into smiles
tmp_fileList = os.listdir('.')
for tmp_file in tmp_fileList:
	tmp_m = re.match('([CH0-9_]+)\.sdf|([CH0-9_]+)_resonance\.sdf|([CHr0-9_]+)\.sdf|([CHr0-9_]+)_resonance\.sdf', tmp_file)
	if tmp_m:
		if tmp_m.group(1) or tmp_m.group(2) :
			output = os.system('E:\\hetanjin\\softwares\\OpenBabel-2.3.72\\babel.exe -isdf ' + tmp_file + ' -osmi ' + tmp_file[0:-4] + '.smi -xnU' + ' > babelLog.txt 2>&1')
		else:
			output = os.system('E:\\hetanjin\\softwares\\OpenBabel-2.3.72\\babel.exe -isdf ' + tmp_file + ' -osmi ' + tmp_file[0:-4] + '.smi -xnUh' + ' > babelLog.txt 2>&1')
		fr = file('babelLog.txt', 'r')
		babelInfo = fr.readlines()
		fr.close()
		for tmp_line in babelInfo:
			if re.search('Warning', tmp_line):
				if tmp_m.group(1) or tmp_m.group(2) :
					output = os.system('E:\\hetanjin\\softwares\\OpenBabel-2.3.72\\babel.exe -isdf ' + tmp_file + ' -osmi ' + tmp_file[0:-4] + '.smi -xni' + ' > babelLog.txt 2>&1')
				else:
					output = os.system('E:\\hetanjin\\softwares\\OpenBabel-2.3.72\\babel.exe -isdf ' + tmp_file + ' -osmi ' + tmp_file[0:-4] + '.smi -xnhi' + ' > babelLog.txt 2>&1')
				fr = file('babelLog.txt', 'r')
				babelInfo2 = fr.readlines()
				fr.close()
				for tmp2_line in babelInfo2:
					if re.search('Warning', tmp2_line):
						print 'babel warning! ' + tmp_file 
						print babelInfo2
						break
				resonanceWarning.append(tmp_file[0:-4])
				break

# read .smi files and write smiles into database excel
tmp_fileList = os.listdir('.')
for tmp_file in tmp_fileList:
	if tmp_file[0:-4] in resonanceWarning:
		print 'Warning! A babel warned stereo ambiguous smiles used!' + tmp_file
		# continue
	tmp_m = re.match('(\w+)_resonance.smi|(\w+).smi', tmp_file)
	if tmp_m:
		fr = file(tmp_file, 'r')
		tmp_smiles = fr.readline()
		tmp_smiles = tmp_smiles.strip()
		fr.close()
		if tmp_m.group(1) == None:
			if tmp_m.group(2) not in smilesDict.keys():
				smilesDict[tmp_m.group(2)] = [tmp_smiles]
			else:
				smilesDict[tmp_m.group(2)][0] = tmp_smiles		
		else:
			if tmp_m.group(1) not in smilesDict.keys():
				smilesDict[tmp_m.group(1)] = ['',tmp_smiles]
			else:
				smilesDict[tmp_m.group(1)].append(tmp_smiles)

for (index_mole, tmp_label) in enumerate(all_labels):
	if tmp_label in smilesDict.keys():
		if len(smilesDict[tmp_label]) == 2 and smilesDict[tmp_label][0] != '':
			tmp_smiles = smilesDict[tmp_label][0] + '\n or \n' + smilesDict[tmp_label][1]
		else:
			tmp_smiles = ''.join(smilesDict[tmp_label])
		tmp_smiles=tmp_smiles.strip()
		shrw.cell(row=2+index_mole, column=2).value = tmp_smiles
	else:
		print 'Error! Smiles is not found for ' + tmp_label 

wbrw.save(dataFile)
